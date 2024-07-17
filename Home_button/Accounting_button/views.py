# Accounting_button/views.py

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import redirect
from django.views.generic import TemplateView, FormView, View
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView as AuthLoginView


class LoginView(AuthLoginView):
    template_name = 'Accounting_button/login.html'

    def get_success_url(self):
        return reverse_lazy('dashboard_redirect')



@method_decorator(login_required, name='dispatch')
class DashboardRedirectView(View):
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.user_type == 'owner':
            return redirect('owner_dashboard')
        elif user.user_type == 'organizer':
            return redirect('organizer_dashboard')
        elif user.user_type == 'executor':
            return redirect('executor_dashboard')
        return redirect('login')

@method_decorator(login_required, name='dispatch')
class OwnerDashboardView(TemplateView):
    template_name = 'Accounting_button/owner_dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.user_type != 'owner':
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class OrganizerDashboardView(TemplateView):
    template_name = 'Accounting_button/organizer_dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.user_type != 'organizer':
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class ExecutorDashboardView(TemplateView):
    template_name = 'Accounting_button/executor_dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.user_type != 'executor':
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)


from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import redirect
from .models import PositionDirectory, TariffDirectory, StaffSchedule



# Декоратор для проверки, что пользователь не является исполнителем
def executor_forbidden(view_func):
    def _wrapped_view(request, *args, **kwargs):
        if request.user.user_type == 'executor':
            return redirect('forbidden')  # Перенаправление на страницу с сообщением о недостаточности прав
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# Декоратор для проверки, что пользователь является собственником
def owner_required(view_func):
    @login_required
    @executor_forbidden
    def _wrapped_view(request, *args, **kwargs):
        if request.user.user_type != 'owner':
            return redirect('forbidden')  # Перенаправление на страницу с сообщением о недостаточности прав
        return view_func(request, *args, **kwargs)
    return _wrapped_view

# Декоратор для проверки, что пользователь является собственником или организатором
def owner_or_organizer_required(view_func):
    @login_required
    @executor_forbidden
    def _wrapped_view(request, *args, **kwargs):
        if request.user.user_type not in ['owner', 'organizer']:
            return redirect('forbidden')  # Перенаправление на страницу с сообщением о недостаточности прав
        return view_func(request, *args, **kwargs)
    return _wrapped_view






# Список всех должностей
@method_decorator(owner_or_organizer_required, name='dispatch')
class PositionDirectoryListView(ListView):
    model = PositionDirectory
    template_name = 'Accounting_button/positions_list.html'

# Детали конкретной должности
@method_decorator(owner_or_organizer_required, name='dispatch')
class PositionDirectoryDetailView(DetailView):
    model = PositionDirectory
    template_name = 'Accounting_button/position_detail.html'

# Создание новой должности (только для собственников)
@method_decorator(owner_required, name='dispatch')
class PositionDirectoryCreateView(CreateView):
    model = PositionDirectory
    template_name = 'Accounting_button/position_form.html'
    fields = ['position', 'comments']
    success_url = reverse_lazy('positions_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Create Position'
        return context

# Редактирование существующей должности (только для собственников)
@method_decorator(owner_required, name='dispatch')
class PositionDirectoryUpdateView(UpdateView):
    model = PositionDirectory
    template_name = 'Accounting_button/position_form.html'
    fields = ['position', 'comments']
    success_url = reverse_lazy('positions_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Edit Position'
        return context

# Удаление существующей должности (только для собственников)
@method_decorator(owner_required, name='dispatch')
class PositionDirectoryDeleteView(DeleteView):
    model = PositionDirectory
    template_name = 'Accounting_button/position_confirm_delete.html'
    success_url = reverse_lazy('positions_list')





# Список всех тарифов
@method_decorator(owner_or_organizer_required, name='dispatch')
class TariffDirectoryListView(ListView):
    model = TariffDirectory
    template_name = 'Accounting_button/tariffs_list.html'  # Исправлено

# Детали конкретного тарифа
@method_decorator(owner_or_organizer_required, name='dispatch')
class TariffDirectoryDetailView(DetailView):
    model = TariffDirectory
    template_name = 'Accounting_button/tariff_detail.html'  # Исправлено

# Создание нового тарифа (только для собственников)
@method_decorator(owner_required, name='dispatch')
class TariffDirectoryCreateView(CreateView):
    model = TariffDirectory
    template_name = 'Accounting_button/tariff_form.html'  # Исправлено
    fields = ['tariff_name', 'cost_per_minute', 'comments']
    success_url = reverse_lazy('tariffs_list')

# Редактирование существующего тарифа (только для собственников)
@method_decorator(owner_required, name='dispatch')
class TariffDirectoryUpdateView(UpdateView):
    model = TariffDirectory
    template_name = 'Accounting_button/tariff_form.html'  # Исправлено
    fields = ['tariff_name', 'cost_per_minute', 'comments']
    success_url = reverse_lazy('tariffs_list')

# Удаление существующего тарифа (только для собственников)
@method_decorator(owner_required, name='dispatch')
class TariffDirectoryDeleteView(DeleteView):
    model = TariffDirectory
    template_name = 'Accounting_button/tariff_confirm_delete.html'  # Исправлено
    success_url = reverse_lazy('tariffs_list')




# Список всего штатного расписания
from django.shortcuts import render
from django.views.generic import ListView
from .models import StaffSchedule

@method_decorator(owner_or_organizer_required, name='dispatch')
class StaffScheduleListView(ListView):
    model = StaffSchedule
    template_name = 'Accounting_button/staff_schedule_list.html'  # Исправлено

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        total_quantity = sum(schedule.quantity for schedule in self.object_list)
        total_working_time_minutes = sum(schedule.total_working_time for schedule in self.object_list)
        total_salary_fund = sum(schedule.total_salary_fund for schedule in self.object_list)
        total_working_time_hours = total_working_time_minutes // 60
        context['total_quantity'] = total_quantity
        context['total_working_time_minutes'] = total_working_time_minutes
        context['total_working_time_hours'] = total_working_time_hours
        context['total_salary_fund'] = total_salary_fund
        return context


# Детали конкретного штатного расписания
@method_decorator(owner_or_organizer_required, name='dispatch')
class StaffScheduleDetailView(DetailView):
    model = StaffSchedule
    template_name = 'Accounting_button/staff_schedule_detail.html'  # Исправлено

# Создание нового штатного расписания (только для собственников)
@method_decorator(owner_required, name='dispatch')
class StaffScheduleCreateView(CreateView):
    model = StaffSchedule
    template_name = 'Accounting_button/staff_schedule_form.html'  # Исправлено
    fields = ['position', 'tariff', 'quantity', 'norm_time_per_month']
    success_url = reverse_lazy('staff_schedule_list')

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.total_working_time = obj.quantity * obj.norm_time_per_month
        obj.total_salary_fund = obj.total_working_time * obj.tariff.cost_per_minute
        obj.save()
        return super().form_valid(form)

# Редактирование существующего штатного расписания (только для собственников)
@method_decorator(owner_required, name='dispatch')
class StaffScheduleUpdateView(UpdateView):
    model = StaffSchedule
    template_name = 'Accounting_button/staff_schedule_form.html'  # Исправлено
    fields = ['position', 'tariff', 'quantity', 'norm_time_per_month']
    success_url = reverse_lazy('staff_schedule_list')

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.total_working_time = obj.quantity * obj.norm_time_per_month
        obj.total_salary_fund = obj.total_working_time * obj.tariff.cost_per_minute
        obj.save()
        return super().form_valid(form)

# Удаление существующего штатного расписания (только для собственников)
@method_decorator(owner_required, name='dispatch')
class StaffScheduleDeleteView(DeleteView):
    model = StaffSchedule
    template_name = 'Accounting_button/staff_schedule_confirm_delete.html'  # Исправлено
    success_url = reverse_lazy('staff_schedule_list')

from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.utils.decorators import method_decorator
from .models import OrganizerPositionDirectory
from .decorators import owner_required, owner_or_organizer_required  # Импорт декораторов

# Представление для списка должностей
@method_decorator(owner_or_organizer_required, name='dispatch')
class OrganizerPositionDirectoryListView(ListView):
    model = OrganizerPositionDirectory
    template_name = 'Accounting_button/organizer_positions/organizer_position_list.html'
    context_object_name = 'object_list'

# Представление для создания новой должности (только для собственников)
@method_decorator(owner_required, name='dispatch')
class OrganizerPositionDirectoryCreateView(CreateView):
    model = OrganizerPositionDirectory
    template_name = 'Accounting_button/organizer_positions/organizer_position_form.html'
    fields = ['position', 'comments']
    success_url = reverse_lazy('organizer_positions_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Создание должности'
        return context

# Представление для деталей должности
@method_decorator(owner_or_organizer_required, name='dispatch')
class OrganizerPositionDirectoryDetailView(DetailView):
    model = OrganizerPositionDirectory
    template_name = 'Accounting_button/organizer_positions/organizer_position_detail.html'

# Представление для редактирования должности (только для собственников)
@method_decorator(owner_required, name='dispatch')
class OrganizerPositionDirectoryUpdateView(UpdateView):
    model = OrganizerPositionDirectory
    template_name = 'Accounting_button/organizer_positions/organizer_position_form.html'
    fields = ['position', 'comments']
    success_url = reverse_lazy('organizer_positions_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Редактирование должности'
        return context

# Представление для удаления должности (только для собственников)
@method_decorator(owner_required, name='dispatch')
class OrganizerPositionDirectoryDeleteView(DeleteView):
    model = OrganizerPositionDirectory
    template_name = 'Accounting_button/organizer_positions/organizer_position_confirm_delete.html'
    success_url = reverse_lazy('organizer_positions_list')

from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from .models import OrganizerTariff, OrganizerPositionDirectory
from .decorators import owner_required, owner_or_organizer_required

# Представление для списка тарифов организаторов
@method_decorator(owner_or_organizer_required, name='dispatch')
class OrganizerTariffListView(ListView):
    model = OrganizerTariff
    template_name = 'Accounting_button/organizer_tariffs/organizer_tariff_list.html'
    context_object_name = 'tariffs'

# Представление для создания нового тарифа (только для собственников)
@method_decorator(owner_required, name='dispatch')
class OrganizerTariffCreateView(CreateView):
    model = OrganizerTariff
    template_name = 'Accounting_button/organizer_tariffs/organizer_tariff_form.html'
    fields = ['position', 'rate', 'base']
    success_url = reverse_lazy('organizer_tariffs_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Создание тарифа'
        return context

# Представление для деталей тарифа
@method_decorator(owner_or_organizer_required, name='dispatch')
class OrganizerTariffDetailView(DetailView):
    model = OrganizerTariff
    template_name = 'Accounting_button/organizer_tariffs/organizer_tariff_detail.html'

# Представление для редактирования тарифа (только для собственников)
@method_decorator(owner_required, name='dispatch')
class OrganizerTariffUpdateView(UpdateView):
    model = OrganizerTariff
    template_name = 'Accounting_button/organizer_tariffs/organizer_tariff_form.html'
    fields = ['position', 'rate', 'base']
    success_url = reverse_lazy('organizer_tariffs_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Редактирование тарифа'
        return context

# Представление для удаления тарифа (только для собственников)
@method_decorator(owner_required, name='dispatch')
class OrganizerTariffDeleteView(DeleteView):
    model = OrganizerTariff
    template_name = 'Accounting_button/organizer_tariffs/organizer_tariff_confirm_delete.html'
    success_url = reverse_lazy('organizer_tariffs_list')



from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from .models import TaxSystem
from .decorators import owner_required, owner_or_organizer_required

# Представление для списка систем налогообложения
@method_decorator(owner_or_organizer_required, name='dispatch')
class TaxSystemListView(ListView):
    model = TaxSystem
    template_name = 'Accounting_button/tax_systems/tax_system_list.html'
    context_object_name = 'tax_systems'

# Представление для создания новой системы налогообложения (только для собственников)
@method_decorator(owner_required, name='dispatch')
class TaxSystemCreateView(CreateView):
    model = TaxSystem
    template_name = 'Accounting_button/tax_systems/tax_system_form.html'
    fields = ['name', 'comments']
    success_url = reverse_lazy('tax_systems_list')

    def form_valid(self, form):
        form.instance.author = self.request.user  # Установить автора
        form.instance.author_name = self.request.user.get_full_name() or self.request.user.username  # Сохранить имя автора
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Создание системы налогообложения'
        return context

# Представление для деталей системы налогообложения
@method_decorator(owner_or_organizer_required, name='dispatch')
class TaxSystemDetailView(DetailView):
    model = TaxSystem
    template_name = 'Accounting_button/tax_systems/tax_system_detail.html'

# Представление для редактирования системы налогообложения (только для собственников)
@method_decorator(owner_required, name='dispatch')
class TaxSystemUpdateView(UpdateView):
    model = TaxSystem
    template_name = 'Accounting_button/tax_systems/tax_system_form.html'
    fields = ['name', 'comments']
    success_url = reverse_lazy('tax_systems_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['view'] = self
        context['title'] = 'Редактирование системы налогообложения'
        return context

# Представление для удаления системы налогообложения (только для собственников)
@method_decorator(owner_required, name='dispatch')
class TaxSystemDeleteView(DeleteView):
    model = TaxSystem
    template_name = 'Accounting_button/tax_systems/tax_system_confirm_delete.html'
    success_url = reverse_lazy('tax_systems_list')


from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .models import Client
from .forms import ClientForm
from .decorators import owner_required, owner_or_organizer_required, executor_forbidden


from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from .models import Client
from .forms import ClientForm




# Представление для отображения списка клиентов
@method_decorator(login_required, name='dispatch')
class ClientListView(ListView):
    model = Client  # Указываем модель для списка
    template_name = 'Accounting_button/Client_list/client_list.html'
    context_object_name = 'clients'  # Имя переменной в контексте шаблона

    def get_queryset(self):
        # Организаторы и исполнители видят только тех клиентов, которые не скрыты в списке
        if self.request.user.user_type in ['organizer', 'executor']:
            return Client.objects.filter(hide_in_list=False)
        # Собственники видят всех клиентов
        return Client.objects.all()

# Представление для отображения деталей клиента
@method_decorator(login_required, name='dispatch')
class ClientDetailView(DetailView):
    model = Client  # Указываем модель для деталей
    template_name = 'Accounting_button/Client_list/client_detail.html'
    context_object_name = 'client'  # Имя переменной в контексте шаблона

# Представление для создания нового клиента (доступно для собственников и организаторов)
@method_decorator(owner_or_organizer_required, name='dispatch')
class ClientCreateView(CreateView):
    model = Client  # Указываем модель для создания
    form_class = ClientForm  # Указываем форму для создания клиента
    template_name = 'Accounting_button/Client_list/client_form.html'
    success_url = reverse_lazy('client_list')  # URL для перенаправления после успешного создания

    def form_valid(self, form):
        form.instance.author = self.request.user  # Устанавливаем автора записи
        return super().form_valid(form)

# Представление для редактирования клиента (доступно для собственников, организаторов и исполнителей)
@method_decorator(login_required, name='dispatch')
class ClientUpdateView(UpdateView):
    model = Client  # Указываем модель для редактирования
    form_class = ClientForm  # Указываем форму для редактирования клиента
    template_name = 'Accounting_button/Client_list/client_form.html'
    success_url = reverse_lazy('client_list')  # URL для перенаправления после успешного редактирования

    def get_form(self):
        form = super().get_form()
        # Ограничиваем доступ для редактирования поля short_name для организаторов и исполнителей
        if self.request.user.user_type != 'owner':
            form.fields['short_name'].disabled = True
        return form

# Представление для удаления клиента (доступно только для собственников)
@method_decorator(owner_required, name='dispatch')
class ClientDeleteView(DeleteView):
    model = Client  # Указываем модель для удаления
    template_name = 'Accounting_button/Client_list/client_confirm_delete.html'
    success_url = reverse_lazy('client_list')  # URL для перенаправления после успешного удаления


from django.http import HttpResponse
import openpyxl
from io import BytesIO
from .models import Client

def export_clients(request):
    # Создаем новый workbook и активный лист
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Clients'

    # Заголовки колонок
    columns = [
        'ID', 'Short Name', 'Full Name', 'Contract Price', 
        'Contract Number and Date', 'INN', 'Tax System', 
        'Nomenclature Units', 'Activity Types', 'Contact Name', 
        'Phone Number', 'Email', 'Postal Address', 'Comment', 
        'Hide In List', 'Author', 'Author Name'
    ]
    row_num = 1

    # Заполняем заголовки
    for col_num, column_title in enumerate(columns, 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Заполняем данные клиентов
    for client in Client.objects.all():
        row_num += 1
        row = [
            client.id, client.short_name, client.full_name, client.contract_price, 
            client.contract_number_date, client.inn, client.tax_system.name, 
            client.nomenclature_units, client.activity_types, client.contact_name, 
            client.phone_number, client.email, client.postal_address, client.comment, 
            client.hide_in_list, client.author.get_full_name() if client.author else '', client.author_name
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Сохраняем workbook в байтовый объект
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Возвращаем ответ с файлом Excel
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=clients.xlsx'
    return response


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import WorkTypeGroup
from .forms import WorkTypeGroupForm
from .decorators import owner_required, owner_or_organizer_required

# Представление для отображения списка всех групп видов работ
@login_required
def worktypegroup_list(request):
    groups = WorkTypeGroup.objects.all()
    return render(request, 'Accounting_button/WorkTypeGroup/list.html', {'groups': groups})

# Представление для создания новой группы видов работ
@owner_or_organizer_required
def worktypegroup_create(request):
    if request.method == 'POST':
        form = WorkTypeGroupForm(request.POST)
        if form.is_valid():
            worktypegroup = form.save(commit=False)
            worktypegroup.author = request.user
            worktypegroup.save()
            return redirect('worktypegroup_list')
    else:
        form = WorkTypeGroupForm()
    return render(request, 'Accounting_button/WorkTypeGroup/form.html', {'form': form})

# Представление для редактирования существующей группы видов работ
@owner_or_organizer_required
def worktypegroup_edit(request, pk):
    worktypegroup = get_object_or_404(WorkTypeGroup, pk=pk)
    if request.method == 'POST':
        form = WorkTypeGroupForm(request.POST, instance=worktypegroup)
        if form.is_valid():
            form.save()
            return redirect('worktypegroup_list')
    else:
        form = WorkTypeGroupForm(instance=worktypegroup)
    return render(request, 'Accounting_button/WorkTypeGroup/form.html', {'form': form})

# Представление для удаления группы видов работ
@owner_required
def worktypegroup_delete(request, pk):
    worktypegroup = get_object_or_404(WorkTypeGroup, pk=pk)
    if request.method == 'POST':
        worktypegroup.delete()
        return redirect('worktypegroup_list')
    return render(request, 'Accounting_button/WorkTypeGroup/confirm_delete.html', {'worktypegroup': worktypegroup})


from django.shortcuts import render, get_object_or_404, redirect
from .models import WorkType
from .forms import WorkTypeForm
from .decorators import unique_owner_required, owner_or_organizer_required, unique_executor_required

@login_required
def worktype_list(request):
    """
    Представление для отображения списка всех видов работ.
    Пользователь с типом 'executor' видит только те записи, у которых нет галки в поле hide_in_list.
    """
    if request.user.user_type == 'executor':
        worktypes = WorkType.objects.filter(hide_in_list=False)
    else:
        worktypes = WorkType.objects.all()
    return render(request, 'Accounting_button/WorkTypes/list.html', {'worktypes': worktypes})

@owner_or_organizer_required
def worktype_create(request):
    """
    Представление для создания нового вида работ.
    Доступно для пользователей с типом 'owner' и 'organizer'.
    """
    if request.method == 'POST':
        form = WorkTypeForm(request.POST)
        if form.is_valid():
            worktype = form.save(commit=False)
            worktype.author = request.user
            worktype.save()
            return redirect('worktype_list')
    else:
        form = WorkTypeForm()
    return render(request, 'Accounting_button/WorkTypes/form.html', {'form': form})

@owner_or_organizer_required
def worktype_edit(request, pk):
    """
    Представление для редактирования существующего вида работ.
    Доступно для пользователей с типом 'owner' и 'organizer'.
    """
    worktype = get_object_or_404(WorkType, pk=pk)
    if request.method == 'POST':
        form = WorkTypeForm(request.POST, instance=worktype)
        if form.is_valid():
            form.save()
            return redirect('worktype_list')
    else:
        form = WorkTypeForm(instance=worktype)
    return render(request, 'Accounting_button/WorkTypes/form.html', {'form': form})

@unique_owner_required
def worktype_delete(request, pk):
    """
    Представление для удаления вида работ.
    Доступно только для пользователей с типом 'owner'.
    Отображает страницу подтверждения перед удалением.
    """
    worktype = get_object_or_404(WorkType, pk=pk)
    if request.method == 'POST':
        worktype.delete()
        return redirect('worktype_list')
    return render(request, 'Accounting_button/WorkTypes/confirm_delete.html', {'worktype': worktype})
